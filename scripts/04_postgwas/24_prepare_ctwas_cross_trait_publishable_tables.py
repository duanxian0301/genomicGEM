from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


PAIR_DEFS = [
    ("F1", "AD"),
    ("F2", "AD"),
    ("F1", "PD"),
    ("F2", "PD"),
    ("F1", "LBD"),
    ("F2", "LBD"),
]

TRAIT_TO_RELEVANT_PAIRS = {
    "F1": ["F1_AD", "F1_PD", "F1_LBD"],
    "F2": ["F2_AD", "F2_PD", "F2_LBD"],
    "AD": ["F1_AD", "F2_AD"],
    "PD": ["F1_PD", "F2_PD"],
    "LBD": ["F1_LBD", "F2_LBD"],
}


@dataclass
class Region:
    chrom: int
    start: int
    end: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary-dir", required=True)
    parser.add_argument("--existing-supp-xlsx", required=True)
    parser.add_argument("--out-xlsx", required=True)
    return parser.parse_args()


def parse_region_id(region_id: str) -> Region:
    chrom, start, end = region_id.split("_")
    return Region(int(chrom), int(start), int(end))


def interval_overlaps(chrom: int, start: int, end: int, other_chrom: int, other_start: int, other_end: int) -> bool:
    return chrom == other_chrom and not (end < other_start or start > other_end)


def infer_coloc_support(pp_h4: float, interpretation: str) -> str | None:
    if pd.notna(interpretation):
        if str(interpretation) == "Strong_H4":
            return "Strong_H4"
        if str(interpretation) == "Moderate_H4":
            return "Moderate_H4"
    if pd.isna(pp_h4):
        return None
    if pp_h4 >= 0.8:
        return "Strong_H4"
    if pp_h4 >= 0.5:
        return "Moderate_H4"
    return None


def infer_pwcoco_support(h4_uncond: float, h4_best: float) -> str | None:
    vals = [x for x in [h4_uncond, h4_best] if pd.notna(x)]
    if not vals:
        return None
    best = max(vals)
    if best >= 0.8:
        return "Strong_H4"
    if best >= 0.5:
        return "Moderate_H4"
    return None


def join_unique(values: Iterable[str]) -> str:
    tokens: set[str] = set()
    for v in values:
        if pd.isna(v):
            continue
        sval = str(v)
        if sval == "" or sval == "nan":
            continue
        for tok in sval.split(";"):
            tok = tok.strip()
            if tok:
                tokens.add(tok)
    clean = sorted(tokens)
    return "; ".join(clean)


def flatten_gene_support(gene_df: pd.DataFrame, pleio_df: pd.DataFrame, coloc_df: pd.DataFrame, pwcoco_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for rec in gene_df.to_dict("records"):
        relevant_pairs = TRAIT_TO_RELEVANT_PAIRS.get(rec["trait"], [])
        conj_pairs = []
        coloc_pairs = []
        pwcoco_pairs = []
        for pair in relevant_pairs:
            sub_pleio = pleio_df[pleio_df["pair"] == pair]
            if not sub_pleio.empty:
                hit = sub_pleio.apply(
                    lambda x: int(rec["chr"]) == int(x["chrnum"]) and int(rec["region_start"]) <= int(x["chrpos"]) <= int(rec["region_end"]),
                    axis=1,
                )
                if hit.any():
                    conj_pairs.append(pair)

            sub_coloc = coloc_df[coloc_df["pair"] == pair]
            if not sub_coloc.empty:
                hit = sub_coloc.apply(
                    lambda x: interval_overlaps(
                        int(rec["chr"]), int(rec["region_start"]), int(rec["region_end"]),
                        int(x["chrnum"]), int(x["region_start"]), int(x["region_end"])
                    ),
                    axis=1,
                )
                for _, crow in sub_coloc.loc[hit].iterrows():
                    support = infer_coloc_support(crow["PP.H4.abf"], crow["coloc_interpretation"])
                    if support:
                        coloc_pairs.append(f"{pair}:{support}")

            sub_pwcoco = pwcoco_df[pwcoco_df["pair"] == pair]
            if not sub_pwcoco.empty:
                hit = sub_pwcoco.apply(
                    lambda x: interval_overlaps(
                        int(rec["chr"]), int(rec["region_start"]), int(rec["region_end"]),
                        int(x["chrnum"]), int(x["region_start"]), int(x["region_end"])
                    ),
                    axis=1,
                )
                for _, prow in sub_pwcoco.loc[hit].iterrows():
                    support = infer_pwcoco_support(prow["H4_uncond"], prow["H4_best_cond"])
                    if support:
                        pwcoco_pairs.append(f"{pair}:{support}")

        rec["conjfdr_pair_overlap"] = join_unique(conj_pairs)
        rec["coloc_support_pairs"] = join_unique(coloc_pairs)
        rec["pwcoco_support_pairs"] = join_unique(pwcoco_pairs)
        rec["external_support_any"] = any([rec["conjfdr_pair_overlap"], rec["coloc_support_pairs"], rec["pwcoco_support_pairs"]])
        rows.append(rec)
    return pd.DataFrame(rows)


def build_pairwise_table(gene_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict] = []
    for left, right in PAIR_DEFS:
        pair_name = f"{left}_vs_{right}"
        left_df = gene_df[gene_df["trait"] == left].copy()
        right_df = gene_df[gene_df["trait"] == right].copy()

        if right_df.empty:
            rows.append(
                {
                    "pair": pair_name,
                    "gene_symbol": pd.NA,
                    "classification": "unavailable_failed_run",
                    "left_best_tissue": pd.NA,
                    "right_best_tissue": pd.NA,
                    "left_max_pip": pd.NA,
                    "right_max_pip": pd.NA,
                    "left_min_FDR": pd.NA,
                    "right_min_FDR": pd.NA,
                    "left_priority_label": pd.NA,
                    "right_priority_label": pd.NA,
                    "left_external_support": pd.NA,
                    "right_external_support": pd.NA,
                    "pair_note": f"{right} run failed; no stable pairwise comparison available."
                }
            )
            continue

        merged = left_df.merge(
            right_df,
            on="gene_symbol",
            how="outer",
            suffixes=(f"_{left}", f"_{right}")
        )

        for rec in merged.to_dict("records"):
            left_prior = bool(rec.get(f"prioritized_any_{left}", False)) if pd.notna(rec.get(f"prioritized_any_{left}")) else False
            right_prior = bool(rec.get(f"prioritized_any_{right}", False)) if pd.notna(rec.get(f"prioritized_any_{right}")) else False
            if left_prior and right_prior:
                classification = "shared_prioritized"
            elif left_prior and not right_prior:
                classification = f"{left}_biased"
            elif right_prior and not left_prior:
                classification = f"{right}_biased"
            else:
                classification = "not_prioritized"

            if classification == "not_prioritized":
                continue

            rows.append(
                {
                    "pair": pair_name,
                    "gene_symbol": rec["gene_symbol"],
                    "classification": classification,
                    "left_best_tissue": rec.get(f"best_tissue_{left}"),
                    "right_best_tissue": rec.get(f"best_tissue_{right}"),
                    "left_max_pip": rec.get(f"max_pip_{left}"),
                    "right_max_pip": rec.get(f"max_pip_{right}"),
                    "left_min_FDR": rec.get(f"min_FDR_{left}"),
                    "right_min_FDR": rec.get(f"min_FDR_{right}"),
                    "left_priority_label": rec.get(f"priority_label_{left}"),
                    "right_priority_label": rec.get(f"priority_label_{right}"),
                    "left_external_support": join_unique([
                        rec.get(f"conjfdr_pair_overlap_{left}"),
                        rec.get(f"coloc_support_pairs_{left}"),
                        rec.get(f"pwcoco_support_pairs_{left}"),
                    ]),
                    "right_external_support": join_unique([
                        rec.get(f"conjfdr_pair_overlap_{right}"),
                        rec.get(f"coloc_support_pairs_{right}"),
                        rec.get(f"pwcoco_support_pairs_{right}"),
                    ]),
                    "pair_note": ""
                }
            )
    return pd.DataFrame(rows)


def build_candidate_tiers(gene_df: pd.DataFrame, pair_df: pd.DataFrame) -> pd.DataFrame:
    pair_shared = pair_df[pair_df["classification"] == "shared_prioritized"].groupby("gene_symbol")["pair"].apply(lambda s: "; ".join(sorted(set(s)))).to_dict()
    pair_biased = pair_df[pair_df["classification"].str.contains("_biased", na=False)].groupby("gene_symbol")["pair"].apply(lambda s: "; ".join(sorted(set(s)))).to_dict()

    rows: list[dict] = []
    for gene, sub in gene_df.groupby("gene_symbol"):
        sub = sub.sort_values(["max_pip", "min_FDR"], ascending=[False, True])
        traits_prioritized = sub.loc[sub["prioritized_any"], "trait"].tolist()
        external_supports = []
        for col in ["conjfdr_pair_overlap", "coloc_support_pairs", "pwcoco_support_pairs"]:
            external_supports.extend([v for v in sub[col].tolist() if pd.notna(v) and str(v) != ""])
        external_support_any = len(external_supports) > 0
        n_high = int((sub["priority_label"] == "high_priority").sum())
        n_mod = int((sub["priority_label"] == "moderate_priority").sum())
        shared_pairs = pair_shared.get(gene, "")
        biased_pairs = pair_biased.get(gene, "")

        if n_high > 0 and external_support_any and (shared_pairs or len(set(traits_prioritized)) >= 2):
            tier = "A"
            rationale = "cTWAS-supported with external pleiotropy/colocalization support and cross-trait convergence"
        elif n_high > 0 or (n_mod > 0 and external_support_any):
            tier = "B"
            rationale = "gene-level posterior support present, but multilayer convergence is incomplete"
        else:
            tier = "C"
            rationale = "exploratory support only; retain in supplementary material"

        top = sub.iloc[0]
        rows.append(
            {
                "tier": tier,
                "gene_symbol": gene,
                "best_trait": top["trait"],
                "best_tissue": top["best_tissue"],
                "best_region_id": top["best_region_id"],
                "best_max_pip": top["max_pip"],
                "best_min_FDR": top["min_FDR"],
                "traits_prioritized": "; ".join(sorted(set(traits_prioritized))),
                "shared_pairs": shared_pairs,
                "biased_pairs": biased_pairs,
                "conjfdr_pair_overlap": join_unique(sub["conjfdr_pair_overlap"]),
                "coloc_support_pairs": join_unique(sub["coloc_support_pairs"]),
                "pwcoco_support_pairs": join_unique(sub["pwcoco_support_pairs"]),
                "downstream_priority": {
                    "A": "Prioritize for scPagwas / scDRS / KO",
                    "B": "Secondary follow-up candidate",
                    "C": "Supplementary only"
                }[tier],
                "rationale": rationale,
            }
        )
    out = pd.DataFrame(rows)
    tier_order = pd.CategoricalDtype(["A", "B", "C"], ordered=True)
    out["tier"] = out["tier"].astype(tier_order)
    out = out.sort_values(["tier", "best_max_pip", "best_min_FDR"], ascending=[True, False, True]).reset_index(drop=True)
    return out


def write_sheet(ws, df: pd.DataFrame, title_row: str | None = None) -> None:
    start_row = 1
    if title_row:
        ws.cell(row=1, column=1, value=title_row).font = Font(bold=True)
        start_row = 3
    for c, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=c, value=col).font = Font(bold=True)
    for r, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c, val in enumerate(row, start=1):
            if pd.isna(val):
                val = None
            ws.cell(row=r, column=c, value=val)


def main() -> None:
    args = parse_args()
    summary_dir = Path(args.summary_dir)

    gene_df = pd.read_csv(summary_dir / "ctwas_trait_level_gene_results.tsv", sep="\t")
    tissue_df = pd.read_csv(summary_dir / "ctwas_trait_level_tissue_results.tsv", sep="\t")
    variant_df = pd.read_csv(summary_dir / "ctwas_trait_level_variant_results.tsv", sep="\t")
    status_df = pd.read_csv(summary_dir / "ctwas_trait_run_status.tsv", sep="\t")

    xlsx = Path(args.existing_supp_xlsx)
    pleio_df = pd.read_excel(xlsx, sheet_name="pleio_factor_loci")
    coloc_df = pd.read_excel(xlsx, sheet_name="coloc_summary")
    pwcoco_df = pd.read_excel(xlsx, sheet_name="pwcoco_best")

    pwcoco_df = pwcoco_df.merge(
        coloc_df[["region_id", "pair", "chrnum", "region_start", "region_end"]],
        on=["region_id", "pair"],
        how="left"
    )

    gene_df = flatten_gene_support(gene_df, pleio_df, coloc_df, pwcoco_df)

    pair_df = build_pairwise_table(gene_df)
    candidate_df = build_candidate_tiers(gene_df, pair_df)

    analysis_overview = pd.DataFrame(
        [
            {
                "field": "analysis_type",
                "value": "Trait-wise multigroup cTWAS followed by cross-trait cTWAS comparison",
                "note": "Each trait was analyzed with multigroup cTWAS across 9 brain tissues, but factor-disease pairs were not jointly modeled in one cross-trait run."
            },
            {
                "field": "can_be_called_full_M_cTWAS_pairwise_joint_analysis",
                "value": "No",
                "note": "The pairwise F1/F2 vs AD/PD/LBD stage is post hoc comparison of separate trait-wise runs, not a single cross-trait joint model."
            },
            {
                "field": "recommended_methods_name",
                "value": "cTWAS-based cross-trait gene prioritization using trait-wise multigroup cTWAS",
                "note": "This wording keeps the multigroup tissue modeling and the post hoc trait comparison conceptually separate."
            },
            {
                "field": "recommended_results_name",
                "value": "Trait-wise multigroup cTWAS revealed prioritized genes and tissue-context patterns, followed by cross-trait comparison",
                "note": "Avoid describing the factor-disease comparisons themselves as full M-cTWAS joint analysis."
            },
            {
                "field": "successful_traits",
                "value": "F1; F2; AD; PD",
                "note": "Stable ctwas runs completed."
            },
            {
                "field": "failed_or_unstable_trait",
                "value": "LBD",
                "note": "EM parameter estimation failed twice; retain as exploratory / failed."
            },
        ]
    )

    trait_level_tissue = tissue_df.copy()
    trait_level_tissue["context_signal_note"] = "Context-level posterior support summary, not a formal enrichment p-value test."
    trait_level_tissue = trait_level_tissue.sort_values(
        ["trait", "n_prioritized_features", "group_prior"], ascending=[True, False, False]
    )

    trait_level_gene = gene_df.copy().sort_values(
        ["trait", "prioritized_any", "max_pip", "min_FDR"], ascending=[True, False, False, True]
    )

    notes_failed = status_df[status_df["status"] != "success"].copy()
    if notes_failed.empty:
        notes_failed = pd.DataFrame([{"trait": pd.NA, "status": "none", "note": "No failed runs"}])

    top_variant_rows = (
        variant_df.sort_values(["trait", "pip", "FDR"], ascending=[True, False, True])
        .groupby("trait")
        .head(50)
        .reset_index(drop=True)
    )

    out_xlsx = Path(args.out_xlsx)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    sheet_map = [
        ("analysis_overview", analysis_overview, "Methodological classification and naming guidance"),
        ("trait_level_gene_results", trait_level_gene, "Gene-level posterior summaries with external locus support annotations"),
        ("trait_level_tissue_results", trait_level_tissue, "Trait-wise tissue/context posterior summaries"),
        ("pairwise_shared_genes", pair_df, "Cross-trait shared and biased prioritized genes"),
        ("candidate_gene_tiers", candidate_df, "Candidate gene tiers for downstream single-cell / functional follow-up"),
        ("notes_on_failed_or_unstable_runs", notes_failed, "Failed or unstable runs"),
        ("trait_level_variant_results", top_variant_rows, "Top variant-level posterior evidence by trait (top 50 per successful trait)")
    ]

    for sheet_name, df, title in sheet_map:
        ws = wb.create_sheet(title=sheet_name[:31])
        write_sheet(ws, df, title_row=title)

    wb.save(out_xlsx)

    methods_text = (
        "We performed trait-wise multigroup cTWAS using the ctwas summary-statistics framework, "
        "analyzing each successful trait (F1, F2, AD, and PD) separately while jointly modeling nearby SNPs "
        "and genetically predicted expression across nine Broad FUSION GTEx brain tissues in hg19/GRCh37. "
        "For each trait, the multigroup model estimated gene- and variant-level posterior evidence together with "
        "context-aware group_prior and group_prior_var parameters across tissues. We then carried out cTWAS-based "
        "cross-trait gene prioritization by comparing posterior gene evidence, best-supported tissues, and locus overlap "
        "patterns across factor-disease pairs (F1 vs AD, F2 vs AD, F1 vs PD, F2 vs PD, and exploratory comparisons with LBD). "
        "Because factor-disease pairs were not jointly modeled within a single cross-trait cTWAS run, this comparison stage "
        "was not described as a full pairwise M-cTWAS joint analysis. LBD did not yield a stable solution after repeated attempts "
        "and was therefore retained as exploratory only."
    )

    results_text = (
        "Trait-wise multigroup cTWAS prioritized convergent and trait-biased genes across the ALPS latent factors and NDD traits, "
        "with stable solutions obtained for F1, F2, AD, and PD but not for LBD. The resulting posterior signal highlighted genes "
        "with repeated support across multiple brain contexts as well as genes concentrated in specific tissues, especially cerebellar, "
        "cortical, basal ganglia, hippocampal, and hypothalamic contexts. Cross-trait comparison indicated both shared prioritized genes "
        "and factor- or disease-biased patterns, which could then be interpreted alongside pleiotropic loci from conjFDR and supportive "
        "colocalization evidence from coloc and PWCoCo. These findings support a gene-prioritization framework for downstream biological "
        "interpretation, while remaining agnostic to direct causal proof."
    )

    (summary_dir / "ctwas_cross_trait_methods_draft.txt").write_text(methods_text, encoding="utf-8")
    (summary_dir / "ctwas_cross_trait_results_draft.txt").write_text(results_text, encoding="utf-8")

    pair_df.to_csv(summary_dir / "pairwise_shared_genes.csv", index=False)
    candidate_df.to_csv(summary_dir / "candidate_gene_tiers.csv", index=False)

    print(f"Wrote {out_xlsx}")


if __name__ == "__main__":
    main()
