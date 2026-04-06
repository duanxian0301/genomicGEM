from __future__ import annotations

import csv
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


POSTGWAS = Path(r"D:\文章\GS\postgwas")
SMR_ROOT = POSTGWAS / "06_smr_f1f2_ndd"
SINGLE_ROOT = SMR_ROOT / "single_trait"
SUMMARY_ROOT = SMR_ROOT / "summary_curated"
WORKBOOK = POSTGWAS / "Supplementary_LDSC_MiXeR_pleioFDR_tables.xlsx"
CONVERGENCE_TSV = SMR_ROOT / "summary" / "smr_candidate_convergence.tsv"
CTWAS_SUMMARY = POSTGWAS / "05_mctwas_f1f2_ndd" / "summary" / "ctwas_trait_level_gene_results.tsv"

TRAITS = ["AD", "PD", "LBD", "F1", "F2"]
BRYOIS_CELLS = [
    "Astrocytes",
    "Microglia",
    "Endothelial.cells",
    "Excitatory.neurons",
    "Inhibitory.neurons",
    "OPCs...COPs",
    "Oligodendrocytes",
    "Pericytes",
]
REMOVE_SHEETS = [
    "SMR_trait_summary",
    "SMR_all_results",
    "SMR_candidates",
    "SMR_conjFDR_loci",
    "SMR_bulk_trait_panel",
    "SMR_celltype_trait_panel",
    "SMR_bulk_bonferroni",
    "SMR_bulk_fdr",
    "SMR_bulk_p1e4",
    "SMR_celltype_bonferroni",
    "SMR_celltype_fdr",
    "SMR_celltype_p1e4",
    "cTWAS_primary",
    "cTWAS_secondary",
]
P_THRESHOLD = 1e-4


def parse_smr_file(path: Path, trait: str, panel: str, context: str) -> pd.DataFrame:
    cols = [
        "probeID",
        "ProbeChr",
        "Gene",
        "Probe_bp",
        "topSNP",
        "topSNP_chr",
        "topSNP_bp",
        "b_SMR",
        "se_SMR",
        "p_SMR",
        "p_HEIDI",
        "nsnp_HEIDI",
        "p_GWAS",
        "p_eQTL",
    ]
    df = pd.read_csv(path, sep="\t", usecols=cols)
    if df.empty:
        return df
    df["trait"] = trait
    df["panel"] = panel
    df["context"] = context
    df["source_file"] = str(path)
    df["heidi_supported"] = df["p_HEIDI"].isna() | (df["p_HEIDI"] > 0.01)
    return df


def bh_fdr(pvals: pd.Series) -> pd.Series:
    p = pd.to_numeric(pvals, errors="coerce")
    out = pd.Series([pd.NA] * len(p), index=p.index, dtype="Float64")
    mask = p.notna()
    if not mask.any():
        return out
    vals = p.loc[mask].astype(float)
    order = vals.sort_values().index
    ranked = vals.loc[order].to_numpy()
    n = len(ranked)
    q = ranked * n / pd.Series(range(1, n + 1), dtype=float).to_numpy()
    q = pd.Series(q[::-1]).cummin()[::-1].clip(upper=1.0).to_numpy()
    out.loc[order] = q
    return out


def collect_hits() -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for trait in TRAITS:
        bm_dir = SINGLE_ROOT / trait / "brainmeta_v2_cortex"
        if bm_dir.exists():
            for path in sorted(bm_dir.glob("*.smr")):
                frames.append(parse_smr_file(path, trait, "BrainMeta_v2", "BrainMeta_cortex"))
        gtex_dir = SINGLE_ROOT / trait / "gtex_v8_brain"
        if gtex_dir.exists():
            for path in sorted(gtex_dir.glob("*.smr")):
                context = path.stem.replace(f"{trait}_", "", 1)
                frames.append(parse_smr_file(path, trait, "GTEx_v8", context))
        bryois_dir = SINGLE_ROOT / trait / "bryois2022_celltype"
        if bryois_dir.exists():
            for cell_dir in sorted(p for p in bryois_dir.iterdir() if p.is_dir()):
                for path in sorted(cell_dir.glob("*.smr")):
                    frames.append(parse_smr_file(path, trait, "Bryois2022", cell_dir.name))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def build_overview_df() -> pd.DataFrame:
    rows = []
    for trait in TRAITS:
        bm_n = len(list((SINGLE_ROOT / trait / "brainmeta_v2_cortex").glob("*.smr")))
        gx_n = len(list((SINGLE_ROOT / trait / "gtex_v8_brain").glob("*.smr")))
        rows.append(
            {
                "analysis_module": "BrainMeta_v2_cortex",
                "trait": trait,
                "unit": "chromosome",
                "completed": bm_n,
                "expected": 22,
                "status": "complete" if bm_n == 22 else "incomplete",
                "note": "bulk cortex eQTL SMR",
            }
        )
        rows.append(
            {
                "analysis_module": "GTEx_v8_brain",
                "trait": trait,
                "unit": "tissue",
                "completed": gx_n,
                "expected": 13,
                "status": "complete" if gx_n == 13 else "incomplete",
                "note": "13 brain tissues",
            }
        )
    for cell in BRYOIS_CELLS:
        for trait in TRAITS:
            by_dir = SINGLE_ROOT / trait / "bryois2022_celltype" / cell
            by_n = len(list(by_dir.glob("*.smr"))) if by_dir.exists() else 0
            rows.append(
                {
                    "analysis_module": f"Bryois2022_{cell}",
                    "trait": trait,
                    "unit": "chromosome",
                    "completed": by_n,
                    "expected": 22,
                    "status": "complete" if by_n == 22 else "incomplete",
                    "note": "cell-type eQTL SMR; incomplete cell types are excluded from core result tables",
                }
            )
    return pd.DataFrame(rows)


def build_context_summary(core_hits: pd.DataFrame) -> pd.DataFrame:
    if core_hits.empty:
        return pd.DataFrame()
    grouped = core_hits.groupby(["trait", "panel", "context"], dropna=False)
    rows = []
    for (trait, panel, context), sub in grouped:
        best = sub.sort_values("p_SMR", kind="stable").iloc[0]
        rows.append(
            {
                "trait": trait,
                "panel": panel,
                "context": context,
                "n_hits_p_smr_lt_1e4": len(sub),
                "n_hits_p_smr_lt_1e4_heidi_supported": int(sub["heidi_supported"].sum()),
                "best_gene": best["Gene"],
                "best_probe": best["probeID"],
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
            }
        )
    return pd.DataFrame(rows).sort_values(["trait", "panel", "best_p_SMR"], kind="stable")


def build_gene_summary(core_hits: pd.DataFrame) -> pd.DataFrame:
    if core_hits.empty:
        return pd.DataFrame()
    rows = []
    grouped = core_hits.groupby(["trait", "panel", "Gene"], dropna=False)
    for (trait, panel, gene), sub in grouped:
        best = sub.sort_values("p_SMR", kind="stable").iloc[0]
        rows.append(
            {
                "trait": trait,
                "panel": panel,
                "Gene": gene,
                "best_context": best["context"],
                "best_probe": best["probeID"],
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
                "best_b_SMR": best["b_SMR"],
                "n_contexts_p_smr_lt_1e4": sub["context"].nunique(),
                "n_contexts_heidi_supported": sub.loc[sub["heidi_supported"], "context"].nunique(),
                "any_heidi_supported": bool(sub["heidi_supported"].any()),
            }
        )
    return pd.DataFrame(rows).sort_values(["trait", "best_p_SMR"], kind="stable")


def split_bulk_and_celltype(all_hits: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    bulk = all_hits.loc[all_hits["panel"].isin(["BrainMeta_v2", "GTEx_v8"])].copy()
    cell = all_hits.loc[all_hits["panel"].eq("Bryois2022")].copy()
    return bulk, cell


def add_threshold_columns(hits: pd.DataFrame) -> pd.DataFrame:
    if hits.empty:
        return hits
    out = hits.copy()
    out["n_tests_trait_panel"] = out.groupby(["trait", "panel"])["p_SMR"].transform("size")
    out["bonf_threshold"] = 0.05 / out["n_tests_trait_panel"]
    out["smr_bonferroni"] = out["p_SMR"] < out["bonf_threshold"]
    out["smr_fdr"] = False
    out["p_SMR_FDR"] = pd.NA
    for (trait, panel), idx in out.groupby(["trait", "panel"]).groups.items():
        fdr = bh_fdr(out.loc[idx, "p_SMR"])
        out.loc[idx, "p_SMR_FDR"] = fdr
        out.loc[idx, "smr_fdr"] = pd.to_numeric(fdr, errors="coerce") < 0.05
    out["smr_bonferroni_heidi"] = out["smr_bonferroni"] & out["heidi_supported"]
    out["smr_fdr_heidi"] = out["smr_fdr"] & out["heidi_supported"]
    out["smr_p1e4_heidi"] = (out["p_SMR"] < P_THRESHOLD) & out["heidi_supported"]
    return out


def build_trait_panel_summary(core_hits: pd.DataFrame) -> pd.DataFrame:
    if core_hits.empty:
        return pd.DataFrame()
    rows = []
    grouped = core_hits.groupby(["trait", "panel"], dropna=False)
    for (trait, panel), sub in grouped:
        best = sub.sort_values("p_SMR", kind="stable").iloc[0]
        rows.append(
            {
                "trait": trait,
                "panel": panel,
                "n_hits_p_smr_lt_1e4": len(sub),
                "n_hits_p_smr_lt_1e4_heidi_supported": int(sub["heidi_supported"].sum()),
                "n_unique_genes": sub["Gene"].nunique(),
                "n_unique_contexts": sub["context"].nunique(),
                "best_gene": best["Gene"],
                "best_context": best["context"],
                "best_probe": best["probeID"],
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
            }
        )
    return pd.DataFrame(rows).sort_values(["trait", "panel"], kind="stable")


def subset_smr_results(hits: pd.DataFrame, flag: str) -> pd.DataFrame:
    if hits.empty:
        return pd.DataFrame()
    keep_cols = [
        "trait",
        "panel",
        "context",
        "Gene",
        "probeID",
        "ProbeChr",
        "Probe_bp",
        "topSNP",
        "topSNP_chr",
        "topSNP_bp",
        "b_SMR",
        "se_SMR",
        "p_SMR",
        "p_SMR_FDR",
        "p_HEIDI",
        "nsnp_HEIDI",
        "heidi_supported",
        "n_tests_trait_panel",
        "bonf_threshold",
        "source_file",
    ]
    out = hits.loc[hits[flag], keep_cols].copy()
    return out.sort_values(["trait", "panel", "p_SMR"], kind="stable")


def build_candidate_core() -> pd.DataFrame:
    if not CONVERGENCE_TSV.exists():
        return pd.DataFrame()
    df = pd.read_csv(CONVERGENCE_TSV, sep="\t", low_memory=False)
    keep = df["candidate_tier"].eq("A_high_convergent") | (
        df["candidate_tier"].eq("B_moderate_support")
        & df["heidi_pass"].fillna(False).astype(bool)
        & (pd.to_numeric(df["p_SMR"], errors="coerce") < 1e-6)
    )
    cols = [
        "trait",
        "Gene",
        "chrom",
        "topSNP",
        "p_SMR",
        "p_HEIDI",
        "heidi_pass",
        "best_tissue",
        "best_p",
        "max_pip",
        "ctwas_supported",
        "any_coloc_supported",
        "any_pwcoco_supported",
        "coloc_pairs",
        "candidate_tier",
    ]
    out = df.loc[keep, cols].copy()
    out = out.sort_values(["candidate_tier", "trait", "p_SMR"], kind="stable")
    return out


def load_ctwas_results() -> pd.DataFrame:
    if not CTWAS_SUMMARY.exists():
        return pd.DataFrame()
    df = pd.read_csv(CTWAS_SUMMARY, sep="\t")
    if df.empty:
        return df
    df["min_FDR"] = pd.to_numeric(df["min_FDR"], errors="coerce")
    df["max_pip"] = pd.to_numeric(df["max_pip"], errors="coerce")
    return df


def build_ctwas_primary_secondary(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()
    keep_cols = [
        "trait",
        "gene_symbol",
        "best_feature",
        "best_tissue",
        "best_region_id",
        "chr",
        "best_z",
        "best_p",
        "min_FDR",
        "max_pip",
        "n_tested_contexts",
        "n_prioritized_contexts",
        "n_fdr_contexts",
        "n_highpip_contexts",
        "priority_label",
    ]
    primary = df.loc[df["max_pip"] >= 0.5, keep_cols].copy()
    primary = primary.sort_values(["trait", "max_pip", "min_FDR"], ascending=[True, False, True], kind="stable")
    secondary = df.loc[(df["min_FDR"] < 0.05) & (df["max_pip"] < 0.5), keep_cols].copy()
    secondary = secondary.sort_values(["trait", "min_FDR", "max_pip"], kind="stable")
    return primary, secondary


def export_table(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, sep="\t", index=False)


def refresh_workbook(tables: dict[str, pd.DataFrame]) -> None:
    wb = load_workbook(WORKBOOK)
    for name in REMOVE_SHEETS:
        if name in wb.sheetnames:
            del wb[name]
    wb.save(WORKBOOK)

    with pd.ExcelWriter(WORKBOOK, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for name, df in tables.items():
            safe = df.copy()
            if not safe.empty:
                for col in safe.columns:
                    if pd.api.types.is_bool_dtype(safe[col]):
                        safe[col] = safe[col].astype(str)
            safe.to_excel(writer, sheet_name=name, index=False)


def main() -> None:
    SUMMARY_ROOT.mkdir(parents=True, exist_ok=True)

    all_hits = collect_hits()
    all_hits = add_threshold_columns(all_hits)
    core_hits = all_hits.loc[pd.to_numeric(all_hits["p_SMR"], errors="coerce") < P_THRESHOLD].copy()
    core_hits = core_hits.sort_values(["trait", "panel", "p_SMR"], kind="stable")
    bulk_hits, celltype_hits = split_bulk_and_celltype(all_hits)
    bulk_core_hits, celltype_core_hits = split_bulk_and_celltype(core_hits)

    overview = build_overview_df()
    bulk_trait_panel_summary = build_trait_panel_summary(bulk_core_hits)
    bulk_context_summary = build_context_summary(bulk_core_hits)
    bulk_gene_summary = build_gene_summary(bulk_core_hits)
    celltype_trait_panel_summary = build_trait_panel_summary(celltype_core_hits)
    celltype_context_summary = build_context_summary(celltype_core_hits)
    celltype_gene_summary = build_gene_summary(celltype_core_hits)
    smr_bulk_bonf = subset_smr_results(bulk_hits, "smr_bonferroni_heidi")
    smr_bulk_fdr = subset_smr_results(bulk_hits, "smr_fdr_heidi")
    smr_bulk_p1e4 = subset_smr_results(bulk_hits, "smr_p1e4_heidi")
    smr_cell_bonf = subset_smr_results(celltype_hits, "smr_bonferroni_heidi")
    smr_cell_fdr = subset_smr_results(celltype_hits, "smr_fdr_heidi")
    smr_cell_p1e4 = subset_smr_results(celltype_hits, "smr_p1e4_heidi")
    candidate_core = build_candidate_core()
    ctwas_primary, ctwas_secondary = build_ctwas_primary_secondary(load_ctwas_results())

    export_table(overview, SUMMARY_ROOT / "SMR_overview.tsv")
    export_table(smr_bulk_bonf, SUMMARY_ROOT / "SMR_bulk_bonferroni.tsv")
    export_table(smr_bulk_fdr, SUMMARY_ROOT / "SMR_bulk_fdr.tsv")
    export_table(smr_bulk_p1e4, SUMMARY_ROOT / "SMR_bulk_hit_1e4.tsv")
    export_table(bulk_trait_panel_summary, SUMMARY_ROOT / "SMR_bulk_trait_panel_summary.tsv")
    export_table(bulk_gene_summary, SUMMARY_ROOT / "SMR_bulk_gene_summary.tsv")
    export_table(bulk_context_summary, SUMMARY_ROOT / "SMR_bulk_context_summary.tsv")
    export_table(smr_cell_bonf, SUMMARY_ROOT / "SMR_celltype_bonferroni.tsv")
    export_table(smr_cell_fdr, SUMMARY_ROOT / "SMR_celltype_fdr.tsv")
    export_table(smr_cell_p1e4, SUMMARY_ROOT / "SMR_celltype_hit_1e4.tsv")
    export_table(celltype_trait_panel_summary, SUMMARY_ROOT / "SMR_celltype_trait_panel_summary.tsv")
    export_table(celltype_gene_summary, SUMMARY_ROOT / "SMR_celltype_gene_summary.tsv")
    export_table(celltype_context_summary, SUMMARY_ROOT / "SMR_celltype_context_summary.tsv")
    export_table(candidate_core, SUMMARY_ROOT / "SMR_candidate_core.tsv")
    export_table(ctwas_primary, SUMMARY_ROOT / "cTWAS_primary.tsv")
    export_table(ctwas_secondary, SUMMARY_ROOT / "cTWAS_secondary.tsv")

    tables = {
        "SMR_overview": overview,
        "SMR_bulk_bonferroni": smr_bulk_bonf,
        "SMR_bulk_fdr": smr_bulk_fdr,
        "SMR_bulk_p1e4": smr_bulk_p1e4,
        "SMR_bulk_trait_panel": bulk_trait_panel_summary,
        "SMR_bulk_gene_summary": bulk_gene_summary,
        "SMR_bulk_context_summary": bulk_context_summary,
        "SMR_celltype_bonferroni": smr_cell_bonf,
        "SMR_celltype_fdr": smr_cell_fdr,
        "SMR_celltype_p1e4": smr_cell_p1e4,
        "SMR_celltype_trait_panel": celltype_trait_panel_summary,
        "SMR_celltype_gene_summary": celltype_gene_summary,
        "SMR_celltype_context_summa": celltype_context_summary,
        "SMR_candidate_core": candidate_core,
        "cTWAS_primary": ctwas_primary,
        "cTWAS_secondary": ctwas_secondary,
    }
    refresh_workbook(tables)

    print("all_rows", len(all_hits))
    print("bulk_bonf_rows", len(smr_bulk_bonf))
    print("bulk_fdr_rows", len(smr_bulk_fdr))
    print("bulk_p1e4_rows", len(smr_bulk_p1e4))
    print("celltype_bonf_rows", len(smr_cell_bonf))
    print("celltype_fdr_rows", len(smr_cell_fdr))
    print("celltype_p1e4_rows", len(smr_cell_p1e4))
    print("ctwas_primary_rows", len(ctwas_primary))
    print("ctwas_secondary_rows", len(ctwas_secondary))
    print("candidate_core_rows", len(candidate_core))
    print("workbook", WORKBOOK)


if __name__ == "__main__":
    main()
